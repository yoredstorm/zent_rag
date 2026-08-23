# =============================================================================
# File plugins — CSV, Excel, JSON, PDF (local UPLOAD_DIR)
# =============================================================================
from __future__ import annotations

import csv
import json
import time
from pathlib import Path
from typing import ClassVar

from src.connectors.plugin.base import (
    ConnectionTestResult,
    ConnectorError,
    ConnectorPlugin,
)
from src.connectors.plugin.models import ColumnSchema, SchemaDiscovery, TableSchema
from src.core.config import get_settings


def _resolve_path(object_key: str) -> Path:
    settings = get_settings()
    base = Path(settings.UPLOAD_DIR)
    path = base / object_key
    if not path.exists():
        raise ConnectorError(f"File not found: {object_key}")
    return path


def _infer_type(values: list[str]) -> str:
    non_empty = [v for v in values if v not in (None, "")]
    if not non_empty:
        return "text"
    if all(_is_int(v) for v in non_empty):
        return "integer"
    if all(_is_float(v) for v in non_empty):
        return "number"
    if all(v.lower() in ("true", "false") for v in non_empty):
        return "boolean"
    return "text"


def _is_int(value: str) -> bool:
    try:
        int(value)
        return True
    except ValueError:
        return False


def _is_float(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False


class CsvPlugin(ConnectorPlugin):
    connector_type: ClassVar[str] = "csv"
    capabilities: ClassVar[frozenset[str]] = frozenset({"test", "discover"})

    def _open(self):
        path = _resolve_path(str(self.config.get("object_key") or ""))
        return open(
            path,
            "r",
            encoding=self.config.get("encoding") or "utf-8",
            newline="",
        )

    async def validate(self) -> None:
        if not self.config.get("object_key"):
            raise ConnectorError("csv config requires object_key")
        _resolve_path(str(self.config["object_key"]))

    async def connect(self) -> None:
        fh = self._open()
        self._fh = fh

    async def test_connection(self) -> ConnectionTestResult:
        start = time.perf_counter()
        try:
            await self.validate()
            with self._open() as fh:
                sample = fh.read(64)
            return ConnectionTestResult(
                ok=True,
                latency_ms=(time.perf_counter() - start) * 1000,
                message="ok",
            )
        except ConnectorError as exc:
            return ConnectionTestResult(
                ok=False,
                latency_ms=(time.perf_counter() - start) * 1000,
                message=str(exc),
            )

    async def discover(self) -> SchemaDiscovery:
        await self.validate()
        with self._open() as fh:
            delimiter = self.config.get("delimiter") or ","
            reader = csv.reader(fh, delimiter=delimiter)
            header = next(reader, None)
            if header is None:
                raise ConnectorError("CSV is empty")
            rows = [next(reader, None) for _ in range(100)]
            rows = [r for r in rows if r is not None]
        columns = [
            ColumnSchema(
                name=str(h),
                data_type=_infer_type(
                    [r[i] if i < len(r) else "" for r in rows]
                ),
                nullable=True,
            )
            for i, h in enumerate(header)
        ]
        table = TableSchema(
            name=str(self.config.get("object_key") or "csv").rsplit("/", 1)[-1],
            columns=columns,
            row_count=len(rows) if len(rows) < 100 else None,
        )
        return SchemaDiscovery(tables=[table], source="csv")


class ExcelPlugin(ConnectorPlugin):
    connector_type: ClassVar[str] = "excel"
    capabilities: ClassVar[frozenset[str]] = frozenset({"test", "discover"})

    async def validate(self) -> None:
        if not self.config.get("object_key"):
            raise ConnectorError("excel config requires object_key")
        path = _resolve_path(str(self.config["object_key"]))
        if path.suffix.lower() not in (".xlsx", ".xls"):
            raise ConnectorError("Excel file must be .xlsx or .xls")

    async def connect(self) -> None:
        pass

    async def test_connection(self) -> ConnectionTestResult:
        start = time.perf_counter()
        try:
            await self.validate()
            return ConnectionTestResult(
                ok=True,
                latency_ms=(time.perf_counter() - start) * 1000,
                message="ok",
            )
        except ConnectorError as exc:
            return ConnectionTestResult(
                ok=False,
                latency_ms=(time.perf_counter() - start) * 1000,
                message=str(exc),
            )

    async def discover(self) -> SchemaDiscovery:
        await self.validate()
        from openpyxl import load_workbook

        path = _resolve_path(str(self.config["object_key"]))
        wb = load_workbook(path, read_only=True, data_only=True)
        tables: list[TableSchema] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [list(row) for _, row in zip(range(3), ws.iter_rows(values_only=True))]
            header = [str(c) if c is not None else "" for c in rows[0]] if rows else []
            columns = [
                ColumnSchema(name=h or f"column_{i}", data_type="text")
                for i, h in enumerate(header)
            ]
            tables.append(TableSchema(name=sheet_name, columns=columns))
        wb.close()
        return SchemaDiscovery(tables=tables, source="excel")


class JsonFilePlugin(ConnectorPlugin):
    connector_type: ClassVar[str] = "json_file"
    capabilities: ClassVar[frozenset[str]] = frozenset({"test", "discover"})

    async def validate(self) -> None:
        if not self.config.get("object_key"):
            raise ConnectorError("json_file config requires object_key")
        _resolve_path(str(self.config["object_key"]))

    async def connect(self) -> None:
        pass

    async def test_connection(self) -> ConnectionTestResult:
        start = time.perf_counter()
        try:
            await self.validate()
            data = json.loads(
                _resolve_path(str(self.config["object_key"])).read_text(
                    encoding="utf-8"
                )
            )
            if data is None:
                raise ConnectorError("JSON file is empty")
            return ConnectionTestResult(
                ok=True,
                latency_ms=(time.perf_counter() - start) * 1000,
                message="ok",
            )
        except (ConnectorError, json.JSONDecodeError) as exc:
            return ConnectionTestResult(
                ok=False,
                latency_ms=(time.perf_counter() - start) * 1000,
                message=str(exc),
            )

    async def discover(self) -> SchemaDiscovery:
        await self.validate()
        data = json.loads(
            _resolve_path(str(self.config["object_key"])).read_text(encoding="utf-8")
        )
        items = data if isinstance(data, list) else [data]
        keys: list[str] = []
        types: dict[str, str] = {}
        for item in items[:100]:
            if not isinstance(item, dict):
                continue
            for key, value in item.items():
                if key not in keys:
                    keys.append(key)
                vtype = type(value).__name__
                if key not in types:
                    types[key] = vtype
                elif types[key] != vtype:
                    types[key] = "mixed"
        table = TableSchema(
            name="json",
            columns=[
                ColumnSchema(name=k, data_type=types.get(k, "unknown"))
                for k in keys
            ],
            row_count=len(items) if len(items) < 100 else None,
        )
        return SchemaDiscovery(tables=[table], source="json_file")


class PdfPlugin(ConnectorPlugin):
    connector_type: ClassVar[str] = "pdf"
    capabilities: ClassVar[frozenset[str]] = frozenset({"test"})

    async def validate(self) -> None:
        if not self.config.get("object_key"):
            raise ConnectorError("pdf config requires object_key")
        _resolve_path(str(self.config["object_key"]))

    async def connect(self) -> None:
        pass

    async def test_connection(self) -> ConnectionTestResult:
        start = time.perf_counter()
        try:
            path = _resolve_path(str(self.config.get("object_key") or ""))
            with open(path, "rb") as fh:
                magic = fh.read(5)
            if magic != b"%PDF-":
                raise ConnectorError("Not a valid PDF (missing %PDF header)")
            return ConnectionTestResult(
                ok=True,
                latency_ms=(time.perf_counter() - start) * 1000,
                message="ok",
            )
        except ConnectorError as exc:
            return ConnectionTestResult(
                ok=False,
                latency_ms=(time.perf_counter() - start) * 1000,
                message=str(exc),
            )


def register() -> None:
    from src.connectors.plugin.registry import register_plugin

    register_plugin(CsvPlugin)
    register_plugin(ExcelPlugin)
    register_plugin(JsonFilePlugin)
    register_plugin(PdfPlugin)
