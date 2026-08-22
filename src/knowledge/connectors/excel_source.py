# =============================================================================
# ExcelSourceConnector — hojas → filas → Records (openpyxl, MIT)
# =============================================================================
from __future__ import annotations

from src.knowledge.connectors.base import (
    ConnectorError,
    DiscoveredItem,
    SourceConnector,
)
from src.knowledge.connectors.csv_source import rows_to_records
from src.knowledge.storage import resolve_path


class ExcelSourceConnector(SourceConnector):
    source_type = "excel"
    self_contained = False

    def _sheet(self) -> str | None:
        return self.config.get("sheet") or None

    async def validate(self) -> None:
        object_key = self.config.get("object_key", "")
        if not object_key:
            raise ConnectorError("excel source requires 'object_key' in config")
        path = resolve_path(self.source.organization_id, object_key)
        if not path.exists():
            raise ConnectorError(f"Uploaded file not found: {object_key}")
        try:
            self._load_sheet(path)
        except Exception as exc:
            raise ConnectorError(f"Excel parse failed: {exc}") from exc

    def _load_sheet(self, path) -> list[dict[str, str]]:
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_name = self._sheet() or workbook.sheetnames[0]
        if sheet_name not in workbook.sheetnames:
            raise ConnectorError(f"Sheet '{sheet_name}' not found in workbook")
        sheet = workbook[sheet_name]
        rows_iter = sheet.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter, ())]
        headers = [h or f"col_{i}" for i, h in enumerate(headers)]
        rows: list[dict[str, str]] = []
        for row in rows_iter:
            if row is None or all(v is None for v in row):
                continue
            rows.append(
                {
                    headers[i]: "" if v is None else str(v)
                    for i, v in enumerate(row)
                    if i < len(headers)
                }
            )
        workbook.close()
        return rows

    async def discover(self) -> list[DiscoveredItem]:
        import openpyxl

        path = resolve_path(self.source.organization_id, self.config.get("object_key", ""))
        workbook = openpyxl.load_workbook(path, read_only=True)
        sheets = list(workbook.sheetnames)
        workbook.close()
        return [
            DiscoveredItem(external_id=f"sheet:{s}", label=s, extra={"rows": "?"})
            for s in sheets
        ]

    async def iter_records(self, cursor: dict | None):
        object_key = self.config.get("object_key", "")
        path = resolve_path(self.source.organization_id, object_key)
        if not path.exists():
            raise ConnectorError(f"Uploaded file not found: {object_key}")
        rows = self._load_sheet(path)
        for record in rows_to_records(
            rows, object_key, extra_metadata={"format": "excel", "sheet": self._sheet()}
        ):
            yield record
